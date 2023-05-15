import json

from django.core.exceptions import ValidationError
from django.test import SimpleTestCase
from openpyxl import load_workbook
from openpyxl.cell import Cell

from audit.excel import (
    ExcelExtractionError,
    extract_federal_awards,
    extract_findings_text,
    extract_findings_uniform_guidance,
    extract_corrective_action_plan,
    federal_awards_field_mapping,
    findings_text_field_mapping,
    findings_uniform_guidance_field_mapping,
    corrective_action_field_mapping,
    federal_awards_column_mapping,
    findings_text_column_mapping,
    findings_uniform_guidance_column_mapping,
    corrective_action_column_mapping,
)
from audit.validators import (
    validate_federal_award_json,
    validate_corrective_action_plan_json,
    validate_findings_text_json,
    validate_findings_uniform_guidance_json,
)
from audit.fixtures.excel import (
    FEDERAL_AWARDS_TEMPLATE,
    FEDERAL_AWARDS_ENTRY_FIXTURES,
    CORRECTIVE_ACTION_PLAN_TEMPLATE,
    CORRECTIVE_ACTION_PLAN_ENTRY_FIXTURES,
    FINDINGS_TEXT_TEMPLATE,
    FINDINGS_TEXT_ENTRY_FIXTURES,
    FINDINGS_UNIFORM_GUIDANCE_TEMPLATE,
    FINDINGS_UNIFORM_GUIDANCE_ENTRY_FIXTURES,
)

# Simplest way to create a new copy of simple case rather than getting
# references to things used by other tests:
jsoncopy = lambda v: json.loads(json.dumps(v))


def _set_by_name(workbook, name, value, row_offset=0):
    definition = workbook.defined_names[name]

    sheet_title, cell_coord = next(definition.destinations)

    sheet = workbook[sheet_title]
    cell_range = sheet[cell_coord]

    if isinstance(cell_range, Cell):
        cell_range.value = value
    else:
        cell_range[row_offset][0].value = value


def _add_entry(workbook, row_offset, entry):
    for key, value in entry.items():
        _set_by_name(workbook, key, value, row_offset)


class FederalAwardsExcelTests(SimpleTestCase):
    def test_template_has_named_ranges(self):
        """Test that the FederalAwardsExpended Excel template contains the expected named ranges"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        for name in federal_awards_field_mapping.keys():
            self.assertIsNotNone(workbook.defined_names[name])

        for name in federal_awards_column_mapping:
            self.assertIsNotNone(workbook.defined_names[name])

    # 20230512 HDMS FIXME: Do we want to allow empty xlsx? If the answer is no, then we need to update this test.
    def test_empty_template(self):
        """Test that extraction and validation succeed against the blank template"""
        federal_awards = extract_federal_awards(FEDERAL_AWARDS_TEMPLATE)

        validate_federal_award_json(federal_awards)

    def test_single_federal_awards_entry(self):
        """Test that extraction and validation succeed when there is a single federal awards entry"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_ein", "123456789")
        _set_by_name(workbook, "total_amount_expended", 100)
        _add_entry(workbook, 0, FEDERAL_AWARDS_ENTRY_FIXTURES[0])

        federal_awards = extract_federal_awards(workbook)

        validate_federal_award_json(federal_awards)

    def test_multiple_federal_awards_entries(self):
        """Test that extraction and validation succeed when there are multiple federal awards entries"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_ein", "123456789")
        _set_by_name(workbook, "total_amount_expended", 200)
        for index, entry in enumerate(FEDERAL_AWARDS_ENTRY_FIXTURES):
            _add_entry(workbook, index, entry)

        federal_awards = extract_federal_awards(workbook)

        validate_federal_award_json(federal_awards)

    def test_partial_federal_awards_entry(self):
        """Test that extraction succeeds and validation fails when there are partial federal awards entries"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_ein", "123456789")
        _set_by_name(workbook, "total_amount_expended", 200)

        entry = jsoncopy(FEDERAL_AWARDS_ENTRY_FIXTURES[0])
        del entry["cluster_name"]

        _add_entry(workbook, 0, entry)

        federal_awards = extract_federal_awards(workbook)

        self.assertRaises(ValidationError, validate_federal_award_json, federal_awards)

    def test_federal_awards_type_checking(self):
        """Test that extraction succeeds and validation fails when fields are of the wrong data type"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(workbook, "auditee_ein", "123456789")
        _set_by_name(workbook, "total_amount_expended", 200)
        _add_entry(workbook, 0, FEDERAL_AWARDS_ENTRY_FIXTURES[0])

        test_cases = [
            ("auditee_ein", 123456789),
            ("total_amount_expended", "not a number"),
            ("amount_expended", "not a  number"),
            ("cluster_name", 123),
            ("is_direct", 123),
            ("is_passed", 123),
            ("subrecipient_amount", "not a number"),
            ("program_name", 123),
            ("loan_balance_at_audit_period_end", "not a number"),
            ("is_guaranteed", 123),
            ("is_major", 123),
            ("audit_report_type", 123),
            ("number_of_audit_findings", "not a number"),
            ("federal_agency_prefix", 10),
            ("three_digit_extension", "001"),
            ("state_cluster_name", 123),
        ]

        # validate that each test_case appropriately checks the data type
        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                federal_awards = extract_federal_awards(workbook)

                self.assertRaises(
                    ValidationError, validate_federal_award_json, federal_awards
                )

    def test_federal_awards_custom_formatters(self):
        """Test that custom federal awards field formatters raise the expected error type when data is malformed"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(workbook, "auditee_ein", "123456789")
        _set_by_name(workbook, "amount_expended", 200)
        _add_entry(workbook, 0, FEDERAL_AWARDS_ENTRY_FIXTURES[0])

        test_cases = [
            ("passthrough_name", 0),
            ("passthrough_identifying_number", 0),
        ]

        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                self.assertRaises(
                    ExcelExtractionError, extract_federal_awards, workbook
                )


class CorrectiveActionPlanExcelTests(SimpleTestCase):
    GOOD_UEI = "AAA123456BBB"
    TOO_SHORT_UEI = "AAA123456"
    TOO_MANY_DIGITS_UEI = "AA123456789X"

    def test_template_has_named_ranges(self):
        """Test that the CorrectiveActionPlan Excel template contains the expected named ranges"""
        workbook = load_workbook(CORRECTIVE_ACTION_PLAN_TEMPLATE, data_only=True)

        for name in corrective_action_field_mapping.keys():
            self.assertIsNotNone(workbook.defined_names[name])

        for name in corrective_action_column_mapping:
            self.assertIsNotNone(workbook.defined_names[name])

    def test_single_corrective_action_plan_entry(self):
        """Test that extraction and validation succeed when there is a single corrective action plan entry"""
        workbook = load_workbook(CORRECTIVE_ACTION_PLAN_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", CorrectiveActionPlanExcelTests.GOOD_UEI)

        _add_entry(workbook, 0, CORRECTIVE_ACTION_PLAN_ENTRY_FIXTURES[0])

        corrective_action_plan = extract_corrective_action_plan(workbook)

        validate_corrective_action_plan_json(corrective_action_plan)

    def test_multiple_corrective_action_plan_entries(self):
        """Test that extraction and validation succeed when there are multiple corrective action plan entries"""
        workbook = load_workbook(CORRECTIVE_ACTION_PLAN_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", CorrectiveActionPlanExcelTests.GOOD_UEI)

        for index, entry in enumerate(CORRECTIVE_ACTION_PLAN_ENTRY_FIXTURES):
            _add_entry(workbook, index, entry)

        corrective_action_plan = extract_corrective_action_plan(workbook)

        validate_corrective_action_plan_json(corrective_action_plan)

    def test_partial_corrective_action_plan_entry(self):
        """Test that extraction succeeds and validation fails when there are partial corrective action plan entries"""
        workbook = load_workbook(CORRECTIVE_ACTION_PLAN_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", CorrectiveActionPlanExcelTests.GOOD_UEI)

        entry = jsoncopy(CORRECTIVE_ACTION_PLAN_ENTRY_FIXTURES[0])
        del entry["planned_action"]

        _add_entry(workbook, 0, entry)

        corrective_action_plan = extract_corrective_action_plan(workbook)

        self.assertRaises(
            ValidationError,
            validate_corrective_action_plan_json,
            corrective_action_plan,
        )

    def test_corrective_action_plan_checking(self):
        """Test that extraction succeeds and validation fails when fields are of the wrong data type"""
        workbook = load_workbook(CORRECTIVE_ACTION_PLAN_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(workbook, "auditee_uei", CorrectiveActionPlanExcelTests.GOOD_UEI)
        _add_entry(workbook, 0, CORRECTIVE_ACTION_PLAN_ENTRY_FIXTURES[0])

        test_cases = [
            ("auditee_uei", CorrectiveActionPlanExcelTests.TOO_SHORT_UEI),
            ("contains_chart_or_table", "not a boolean"),
            ("planned_action", 0),
            ("reference_number", 0),
        ]

        # validate that each test_case appropriately checks the data type
        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                corrective_action_plan = extract_corrective_action_plan(workbook)

                self.assertRaises(
                    ValidationError,
                    validate_corrective_action_plan_json,
                    corrective_action_plan,
                )


class FindingsUniformGuidanceExcelTests(SimpleTestCase):
    def test_template_has_named_ranges(self):
        """Test that the FindingsUniformGuidance Excel template contains the expected named ranges"""
        workbook = load_workbook(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE, data_only=True)

        for name in findings_uniform_guidance_field_mapping.keys():
            self.assertIsNotNone(workbook.defined_names[name])

        for name in findings_uniform_guidance_column_mapping:
            self.assertIsNotNone(workbook.defined_names[name])

    def test_empty_template(self):
        """Test that extraction and validation succeed against the blank template"""
        findings = extract_findings_uniform_guidance(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE)

        validate_findings_uniform_guidance_json(findings)

    def test_single_findings_uniform_guidance_entry(self):
        """Test that extraction and validation succeed when there is a single findings uniform guidance entry"""
        workbook = load_workbook(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_ein", "123456789")
        _add_entry(workbook, 0, FINDINGS_UNIFORM_GUIDANCE_ENTRY_FIXTURES[0])

        findings = extract_findings_uniform_guidance(workbook)

        validate_findings_uniform_guidance_json(findings)

    def test_multiple_findings_uniform_guidance_entries(self):
        """Test that extraction and validation succeed when there are multiple findings uniform guidance entries"""
        workbook = load_workbook(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_ein", "123456789")
        for index, entry in enumerate(FINDINGS_UNIFORM_GUIDANCE_ENTRY_FIXTURES):
            _add_entry(workbook, index, entry)

        findings = extract_findings_uniform_guidance(workbook)

        validate_findings_uniform_guidance_json(findings)

    def test_partial_findings_uniform_guidance_entry(self):
        """Test that extraction succeeds and validation fails when there are partial findings uniform guidance entries"""
        workbook = load_workbook(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_ein", "123456789")

        entry = jsoncopy(FINDINGS_UNIFORM_GUIDANCE_ENTRY_FIXTURES[0])
        del entry["reference_number"]

        _add_entry(workbook, 0, entry)

        findings = extract_findings_uniform_guidance(workbook)

        self.assertRaises(
            ValidationError, validate_findings_uniform_guidance_json, findings
        )

    def test_findings_uniform_guidance_checking(self):
        """Test that extraction succeeds and validation fails when fields are of the wrong data type"""
        workbook = load_workbook(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(workbook, "auditee_ein", "123456789")
        _add_entry(workbook, 0, FINDINGS_UNIFORM_GUIDANCE_ENTRY_FIXTURES[0])

        test_cases = [
            ("auditee_ein", 123456789),
            ("reference_number", 0),
            ("program_name", 123),
            ("federal_agency_prefix", 10),
            ("three_digit_extension", "001"),
            ("prior_references", 123),
        ]

        # validate that each test_case appropriately checks the data type
        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                findings = extract_findings_uniform_guidance(workbook)

                self.assertRaises(
                    ValidationError, validate_findings_uniform_guidance_json, findings
                )


class FindingsTextExcelTests(SimpleTestCase):
    def test_template_has_named_ranges(self):
        """Test that the FindingsText Excel template contains the expected named ranges"""
        workbook = load_workbook(FINDINGS_TEXT_TEMPLATE, data_only=True)

        for name in findings_text_field_mapping.keys():
            self.assertIsNotNone(workbook.defined_names[name])

        for name in findings_text_column_mapping:
            self.assertIsNotNone(workbook.defined_names[name])

    def test_empty_template(self):
        """Test that extraction and validation succeed against the blank template"""
        findings = extract_findings_text(FINDINGS_TEXT_TEMPLATE)

        validate_findings_text_json(findings)

    def test_single_findings_text_entry(self):
        """Test that extraction and validation succeed when there is a single findings text entry"""
        workbook = load_workbook(FINDINGS_TEXT_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_ein", "123456789")
        _add_entry(workbook, 0, FINDINGS_TEXT_ENTRY_FIXTURES[0])

        findings = extract_findings_text(workbook)

        validate_findings_text_json(findings)

    def test_multiple_findings_text_entries(self):
        """Test that extraction and validation succeed when there are multiple findings text entries"""
        workbook = load_workbook(FINDINGS_TEXT_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_ein", "123456789")
        for index, entry in enumerate(FINDINGS_TEXT_ENTRY_FIXTURES):
            _add_entry(workbook, index, entry)

        findings = extract_findings_text(workbook)

        validate_findings_text_json(findings)

    def test_partial_findings_text_entry(self):
        """Test that extraction succeeds and validation fails when there are partial findings text entries"""
        workbook = load_workbook(FINDINGS_TEXT_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_ein", "123456789")

        entry = jsoncopy(FINDINGS_TEXT_ENTRY_FIXTURES[0])
        del entry["text_of_finding"]

        _add_entry(workbook, 0, entry)

        findings = extract_findings_text(workbook)

        self.assertRaises(ValidationError, validate_findings_text_json, findings)

    def test_findings_text_checking(self):
        """Test that extraction succeeds and validation fails when fields are of the wrong data type"""
        workbook = load_workbook(FINDINGS_TEXT_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(workbook, "auditee_ein", "123456789")
        _add_entry(workbook, 0, FINDINGS_TEXT_ENTRY_FIXTURES[0])

        test_cases = [
            ("auditee_ein", 123456789),
            ("reference_number", 0),
            ("contains_chart_or_table", "not a boolean"),
            ("text_of_finding", 10.001),
        ]

        # validate that each test_case appropriately checks the data type
        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                findings = extract_findings_text(workbook)

                self.assertRaises(
                    ValidationError, validate_findings_text_json, findings
                )
